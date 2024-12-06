
#? ----------------------------------------------------------------
#? Record
#? ----------------------------------------------------------------
#? 24.10.20 - 전반적으로 손보기
#? 24.10.29 - 삼일 출고 내역 읽어 올 때 '출고일자'가 비어 있으면 스킵하도록 수정함.
#? 24.10.29 - 반품, 입고 내역 "출고내역"에 반영
#? 24.10.31 - 최근 "출고내역" 에는 최근 1년치 데이터만 업데이트 되도록 함


# 0. 기존 all 데이터를 삼일, 다원 각각 df화 한다.
# 1. "출고 내역 업로드" 폴더에 모든 파일을 읽는다.
# 2. 삼일 파일이면, 삼일 컬럼에 맞게 작업. 
# 3. 다원 파일이면, 다원 컬럼에 맞게 작업. 
# 4. 작업방식: 기존 all 에 있는 날짜를 리스트화 하고, 새로운 파일의 날짜들도 리스트업한다
# 5. 여기서 겹치는 것들을 all 데이터에서 삭제한다.
# 6. 두 df를 합쳐서 새로운 all 이 된다. 
# 7. 이 작업을 반복해서, 폴더 안에 모든 파일들 마무리
# 8. 새로운 삼일 all 에서 필요한 컬럼만 추출하고 "출처" = 삼일 을 추가한다.
# 9. 새로운 다원 all 에서 필요한 컬럼만 추출하고 "출처" = 다원 을 추가한다.
# 10. 컬럼 이름을 통일한다. 
# 11. 두 df를 합친다. 
# 12. 온스용 브랜드 들을 추출한다. 
# 13. 전체는 전체 마스터파일에, 온스용은 온스용 마스터파일에 저장한다. 



import os
import pandas as pd
from tkinter import messagebox
import sys
from openpyxl import load_workbook
from datetime import datetime
import tkinter as tk
import time
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2.service_account import Credentials
import json
from datetime import date
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build

# 인증 설정
SERVICE_ACCOUNT_FILE = 'foroncomm-57ce18a35975.json'
SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

# Google Sheets API 인증 및 연결 함수
def connect_to_gsheet(json_keyfile_name, sheet_name,worksheet_index):
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive","https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(json_keyfile_name, scope)
    client = gspread.authorize(creds)
    sheet = client.open(sheet_name).get_worksheet(worksheet_index)  # 3 번째 시트를 엽니다
    return sheet, creds

# 데이터프레임을 Google Sheets에 업데이트하는 함수
def update_gsheet(sheet, df):
    df = df.applymap(lambda x: x.strftime('%Y-%m-%d') if isinstance(x, pd.Timestamp) else x)  # datetime 형식을 문자열로 변환
    sheet.clear()  # 기존 데이터 삭제
    sheet.update([df.columns.values.tolist()] + df.values.tolist())  # 새로운 데이터로 업데이트


# gsheet 에서 정보를 가져오는 함수
# 누락된 코드가 없는지 확인하기 위함
def get_google_sheets_data(spreadsheet_id, sheet_name, range_name):
    """Google Sheets에서 데이터를 가져오는 함수"""
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)
    
    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=range_name
        ).execute()
        values = result.get('values', [])
        
        if not values:
            print('데이터가 없습니다.')
            return []
        else:
            # 데이터프레임으로 변환
            df = pd.DataFrame(values, columns=['상품코드'])
            return df['상품코드'].tolist()
    except Exception as e:
        messagebox.showerror("에러", f"데이터를 가져오는 중 오류가 발생했습니다: {e}")
        return []

# 가져온 내용이랑 비교해서 누락을 찾는 함수
def compare_data_and_find_missing(existing_list, combined_df):
    """기존 데이터 리스트와 엑셀 파일의 [상품코드] 열을 비교하여 누락된 값을 찾는 함수"""
    new_values = combined_df['상품코드'].unique().tolist()
    new_unique_values = [value for value in new_values if value not in existing_list]

    if len(new_unique_values) > 0:
        # 누락된 값이 있는 경우 메시지 박스 표시
        message = "대시보드에 누락된 코드가 있습니다:\n\n"
        for value in new_unique_values:
            message += str(value) + "\n"
        
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("누락된 코드", message)
    else:
        # 누락된 값이 없는 경우
        print("대시보드에 누락된 코드가 없습니다.")
        print("--------------------------------------------------------")







def checkdateformat(excel_file_path,target_sheet_name):
# 엑셀 파일을 열어서 서식을 변경합니다.
    wb = load_workbook(excel_file_path)
    ws = wb[target_sheet_name]

    # "특정열1" 열의 서식을 "날짜"로 변경합니다.
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            # 셀의 값이 문자열인 경우에만 변환 수행
            if isinstance(cell.value, str):
                # yyyy-mm-dd 형식으로 되어있는 문자열을 파싱하여 년, 월, 일로 분리
                year, month, day = map(int, cell.value.split('-'))
                # 년, 월, 일을 "날짜" 형식으로 변환하여 셀에 쓰기
                cell.value = datetime(year, month, day).date()

    # 변경된 내용을 저장합니다.
    wb.save(excel_file_path)


def update_logis_data(df_logis_all, df_new):
    print("--------------------------------------------------------")
    print("삼일 출고내역 가공을 시작합니다")
    print("--------------------------------------------------------")
    df_logis_all.loc[:,'[출고일자]'] = pd.to_datetime(df_logis_all['[출고일자]']).dt.date
    df_new.loc[:,'[출고일자]'] = pd.to_datetime(df_new['[출고일자]']).dt.date
    lll=[]
    # 마지막 행 삭제
    df_new = df_new.iloc[:-1]

    # 새로운 df의 날짜를 가지고 와서, 이 날짜와 겹치는 all 데이터들 삭제
    lll = df_new['[작업일자]'].unique().tolist()
    # 기존 all df에서, new df 와 작업일자가 겹치는 것들 삭제
    df_logis_all = df_logis_all[~df_logis_all['[작업일자]'].isin(lll)]

    # 두 df 합치기
    df_logis_all = pd.concat([df_logis_all,df_new], ignore_index=True)
    print("삼일 출고내역 가공이 완료 되었습니다")
    print("--------------------------------------------------------")
    
    return df_logis_all


def update_dawon_data(df_dawon_all, df_new):
    print("--------------------------------------------------------")
    print("다원 출고내역 가공을 시작합니다")
    print("--------------------------------------------------------")
    lll=[]
    # 마지막 행 삭제
    df_new = df_new.iloc[:-1]

    #?--------------------------------------------------------
    #? 날짜+시간 형식을 날짜만 있는걸로 바꾸기
    #?--------------------------------------------------------
    # 날짜형식으로 변환
    df_dawon_all.loc[:,'ADDDATETIME'] = pd.to_datetime(df_dawon_all['ADDDATETIME']).dt.date
    df_new.loc[:,'ADDDATETIME'] = pd.to_datetime(df_new['ADDDATETIME']).dt.date
    #?--------------------------------------------------------

    # 가져온 all df 의 작업일자들을 리스트로 변환
    lll = df_new['ADDDATETIME'].unique().tolist()

    # 새로운 df에서, 기존 all df 와 작업일자가 겹치는 것들 삭제
    df_dawon_all = df_dawon_all[~df_dawon_all['ADDDATETIME'].isin(lll)]
    
    # 두 df 합치기
    df_dawon_all = pd.concat([df_dawon_all,df_new], ignore_index=True)
    print("다원 출고내역 가공이 완료 되었습니다")
    print("--------------------------------------------------------")
    return df_dawon_all





"""
프로그램 시작

"""



try:
    #!경고 메세지 뜨는거 귀찮으니까 주석처리0
    #!messagebox.showinfo("파일닫기","마스터 파일을 닫아주세요. 제발요")
    print("--------------------------------------------------------")
    print("출고내역 업데이트를 시작합니다.")
    print("--------------------------------------------------------")
    #!----------------------------------------------------------------

    # 로지스 올데이터 엑셀 파일 경로
    excel_file_path = 'ALL DATA_SAMIL csv.csv'
    # 엑셀 파일을 데이터프레임으로 불러오기
    df_logis_all = pd.read_csv(excel_file_path)
    df_logis_all['[출고일자]'] = pd.to_datetime(df_logis_all['[출고일자]'].str.split().str[0], format='%Y-%m-%d')


    # 다원 올데이터 엑셀 파일 경로
    excel_file_path = 'ALL DATA_DAWON csv.csv'
    # 엑셀 파일을 데이터프레임으로 불러오기
    df_dawon_all = pd.read_csv(excel_file_path, parse_dates=['ADDDATETIME'], low_memory=False)



    # 폴더 경로
    folder_path = '출고내역_업로드'


    # 폴더 내의 모든 엑셀 파일에 대해 반복
    for filename in os.listdir(folder_path):
        if filename.endswith(('.xlsx', '.xls')):
            # 엑셀 파일 경로
            excel_file = os.path.join(folder_path, filename)
            # 엑셀 파일 불러오기
            df_new = pd.read_excel(excel_file)

            #삼일여부 체크
            if '[출고일자]' in df_new.columns:
                print("--------------------------------------------------------")
                print("삼일 출고 파일을 발견했습니다.")
                print("--------------------------------------------------------")
                df_logis_all = update_logis_data(df_logis_all, df_new)
                pass
            
            #다원여부 체크
            elif '주문일' in df_new.columns:
                print("--------------------------------------------------------")
                print("다원 출고 파일을 발견했습니다.")
                print("--------------------------------------------------------")
                df_dawon_all = update_dawon_data(df_dawon_all, df_new)
                pass

            else:
                # 둘 다 없으면 메시지 박스를 통해 사용자에게 알리고 프로그램 종료
                messagebox.showerror("에러", "출고파일이 아닌 파일이 있습니다. 확인 해 주세요")
                sys.exit()



    print("--------------------------------------------------------")
    print("가공된 데이터 취합을 시작합니다.")
    print("--------------------------------------------------------")
    # 삼일 저장할 엑셀 파일 경로
    excel_file_path = 'ALL DATA_SAMIL csv.csv'
    df_logis_all.to_csv(excel_file_path, index=False)  


    # 다원 저장할 엑셀 파일 경로
    excel_file_path = 'ALL DATA_DAWON csv.csv'
    df_dawon_all.to_csv(excel_file_path, index=False)  


    print("전체 데이터 업데이트 완료.")
    print("--------------------------------------------------------")



    #?--------------------------------------------------------
    #? 개별 파일 조금씩 다듬기
    #?--------------------------------------------------------
    # 특정 열의 모든 값에 대해 앞 8개 문자를 지우기
    column_to_modify = '[매출처]'
    df_logis_all[column_to_modify] = df_logis_all[column_to_modify].str[7:]


    #?--------------------------------------------------------
    column_to_modify2 = '[브랜드]'
    # 각 셀의 값을 처리하여 ':' 이후의 문자열을 추출하는 함수
    def process_value(cell_value):
        if ':' in cell_value:
            return cell_value.split(':')[1].strip()  # ':' 이후의 문자열 추출 후 공백 제거
        else:
            return cell_value

    # apply 함수를 사용하여 각 셀의 값을 처리하고 결과를 새로운 열에 할당
    df_logis_all[column_to_modify2] = df_logis_all[column_to_modify2].apply(process_value)

    #?--------------------------------------------------------

    #삼일
    columns_to_copy = ['[출고일자]', '[브랜드]', '[상품코드]', '[수량]', '[매출처]']
    df_logis_all = df_logis_all[columns_to_copy]

    # 출고일자가 비어있지 않은 행만 선택
    df_logis_all = df_logis_all[df_logis_all['[출고일자]'].notna()]

    df_logis_all['출처'] = "삼일"
    df_logis_all.columns = ['출고일자','브랜드','상품코드','출고수량','매출처','출처']

    #앞뒤 공백 제거
    df_logis_all['매출처'] = df_logis_all['매출처'].str.strip()



    #다원
    columns_to_copy = ['출고완료일', 'ITEMGROUP', '품목코드(구성품)', '출고완료', '몰명']
    df_dawon_all = df_dawon_all[columns_to_copy]
    df_dawon_all['출처'] = "다원"
    df_dawon_all.columns = ['출고일자','브랜드','상품코드','출고수량','매출처','출처']
    
    df_dawon_all['매출처'] = df_dawon_all['매출처'].fillna('')
    #앞뒤 공백 제거
    df_dawon_all['매출처'] = df_dawon_all['매출처'].str.strip()
    #df_dawon_all = df_dawon_all[df_dawon_all['매출처'] != ""]

    df_dawon_all['매출처'] = df_dawon_all['매출처'].replace("", "세트")
    

    #?--------------------------------------------------------

    df_combined = pd.DataFrame()
    # 두 데이테 프레임 합치기
    df_combined = pd.concat([df_dawon_all,df_logis_all], ignore_index=True)


    print("가공된 데이터 취합이 완료 되었습니다.")
    print("--------------------------------------------------------")


    #?--------------------------------------------------------



    print("반품, 입고 수량을 반영하기 위한 데이터 취합을 진행합니다.")
    print("--------------------------------------------------------")


    #?--------------------------------------------------------
    #?반품, 입고 수량 반영하기 
    #?--------------------------------------------------------

    # 반품입고_업로드 폴더의 파일 읽기
    directory = '반품입고_업로드'
    for filename in os.listdir(directory):
        if filename.endswith('.xls'):
            filepath = os.path.join(directory, filename)
            df = pd.read_excel(filepath)

            # "반품" 파일 처리
            if "반품" in filename:
                filtered_df = df[(df['DAMAGEFLG'] == '정상품') & df['반품확정수량'].notna()]
                filtered_df = filtered_df[['입고실적일', 'ITEMGROUP', '품목코드', '반품확정수량','고객명']]
                filtered_df.columns = ['출고일자', '브랜드', '상품코드', '출고수량','매출처']
                filtered_df['출고수량'] *= -1
                filtered_df['출처'] = '반품'
                #print(filtered_df)
                df_combined = pd.concat([df_combined,filtered_df], ignore_index=True)

            # "입고" 파일 처리
            elif "입고" in filename:                
                filtered_df = df[(df['입고타입'] == '일반입고') & df['NOTES'].fillna('').str.strip().str.contains('반송', case=False, regex=True)]
                filtered_df = filtered_df[['입고실적일', 'ITEMGROUP', '품목코드', '입고수량(낱개)','NOTES']]
                filtered_df.columns = ['출고일자', '브랜드', '상품코드', '출고수량','매출처']
                filtered_df['출고수량'] *= -1
                filtered_df['출처'] = '반송'
                
                #print(filtered_df)
                df_combined = pd.concat([df_combined,filtered_df], ignore_index=True)

            else:
                continue


    print("반품, 입고 수량을 전체 데이터에 취합 완료 했습니다.")
    print("--------------------------------------------------------")



    column_name = '출고일자'
    # 연-월-일 형식의 데이터를 연-월-일로 변경
    df_combined[column_name] = pd.to_datetime(df_combined[column_name]).dt.date

    # 날짜 형식을 문자열로 변환
    df_combined['출고일자'] = df_combined['출고일자'].astype(str)



    print("최근 1년 데이터만 남겨둡니다.")
    print("--------------------------------------------------------")

    from datetime import datetime, timedelta

    # 현재 날짜와 1년 전 날짜 계산
    today = datetime.now()
    one_year_ago = today - timedelta(days=365)

    # 출고일자가 1년 이내인 데이터만 남기기
    df_combined['출고일자'] = pd.to_datetime(df_combined['출고일자'], errors='coerce', format='%Y-%m-%d')

    # 변환한 날짜를 기준으로 1년 이내의 데이터 필터링
    df_combined = df_combined[(df_combined['출고일자'] >= one_year_ago) & (df_combined['출고일자'] <= today)]



    #?--------------------------------------------------------



    print("모든 기본 데이터 정리가 마무리 되었습니다.")
    print("--------------------------------------------------------")



    print("--------------------------------------------------------")
    print("전체 코드용 파일에 업데이트를 시작합니다.")
    print("--------------------------------------------------------")


    print("전체용 대시보드에 누락된 코드가 없는지 확인합니다.")
    print("--------------------------------------------------------")




    #? 구글 스프레드 시트 활용


    # 누락된 코드가 없는지 확인하기
    # 스프레드시트 ID 및 시트 이름 설정
    spreadsheet_id = '1lXP8VSM25lS4m8eLRFa-g9KV8xLr1kg7ARABRXFrKqw'  # INV_master
    sheet_name = 'Sheet3'  # 시트 이름 (시트 이름을 알고 있어야 합니다)
    range_name = f'{sheet_name}!B6:B200'  # 범위 설정 (B6부터 B60까지)

    # 데이터 가져오기
    selected_list = get_google_sheets_data(spreadsheet_id, sheet_name, range_name)
    #print(selected_list)

    try:
        compare_data_and_find_missing(selected_list, df_combined)
    except FileNotFoundError:
        messagebox.showerror("에러", "새로운 엑셀 파일을 찾을 수 없습니다. 파일 경로를 확인해주세요.")
    except Exception as e:
        messagebox.showerror("에러", f"새로운 데이터프레임을 처리하는 중 오류가 발생했습니다: {e}")





    print("업데이트 내용을 전체코드 마스터 파일에 적용합니다. (시간 좀 걸림)")
    print("--------------------------------------------------------")

    # Google Sheets API 설정
    json_keyfile_name = 'foroncomm-57ce18a35975.json'  # JSON 키 파일 경로를 여기에 입력
    sheet_name = 'INV_master'
    worksheet_index = 2  # 세 번째 시트를 선택

    # Google Sheets에 연결
    sheet, creds = connect_to_gsheet(json_keyfile_name, sheet_name, worksheet_index)

    # Google Sheets에 데이터 업데이트
    update_gsheet(sheet, df_combined)







    #!----------------------------------------------------------------
    #! 온스용 제작 
    #!----------------------------------------------------------------
    print("온스 코드용 파일 업데이트를 시작합니다.")
    print("--------------------------------------------------------")

    # 선택할 값들의 리스트
    target_values = ['니심 ', '란시노 ', '아이로 ','에티튜드','조아써 ','웰라 ','니심' , '란시노' , '아이로' , '조아써']

    # 각 값에 대한 불리언 조건을 리스트 내포를 사용하여 생성
    combined_condition = df_combined['브랜드'].isin(target_values)

    # 조건에 해당하는 행들만 선택
    selected_rows = df_combined[combined_condition].copy()


    print("온스용 대시보드에 누락된 코드가 없는지 확인합니다.")
    print("--------------------------------------------------------")


    # 스프레드시트 ID 및 시트 이름 설정
    spreadsheet_id_ons = '1Jjbb2KSjh0LLRQH0VIDKxvWYZUH8TXFWuS5q7RN4474'  # 스프레드시트의 ID
    sheet_name_ons = 'Sheet3'  # 시트 이름 (시트 이름을 알고 있어야 합니다)
    range_name = f'{sheet_name_ons}!B6:B200'  # 범위 설정 (B6부터 B60까지)

    # 데이터 가져오기
    selected_list = get_google_sheets_data(spreadsheet_id_ons, sheet_name_ons, range_name)
    #print(selected_list)

    try:
        compare_data_and_find_missing(selected_list, selected_rows)
    except FileNotFoundError:
        messagebox.showerror("에러", "새로운 엑셀 파일을 찾을 수 없습니다. 파일 경로를 확인해주세요.")
    except Exception as e:
        messagebox.showerror("에러", f"새로운 데이터프레임을 처리하는 중 오류가 발생했습니다: {e}")


    print("업데이트 내용을 전체코드 마스터 파일에 적용합니다. (시간 좀 걸림)")
    print("--------------------------------------------------------")

    # Google Sheets API 설정
    json_keyfile_name = 'foroncomm-57ce18a35975.json'  # JSON 키 파일 경로를 여기에 입력
    sheet_name = 'INV_master_ONNURISTORE'
    worksheet_index = 2  # 세 번째 시트를 선택

    # Google Sheets에 연결
    sheet, creds = connect_to_gsheet(json_keyfile_name, sheet_name, worksheet_index)

    # Google Sheets에 데이터 업데이트
    update_gsheet(sheet, selected_rows)



    print("모든 출고데이터 업데이트가 완료 되었습니다.")
    print("--------------------------------------------------------")



    input("Press Enter to exit...")

except Exception as e:
    print(f"에러 발생: {e}")
    print(f"에러 위치: {e.__traceback__.tb_lineno}")
    print(f"오류 내용: {e.__cause__}")
    input("Press Enter to exit...")