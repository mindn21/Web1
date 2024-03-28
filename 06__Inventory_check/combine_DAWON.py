import os
import pandas as pd
import time
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
# 0. 기존 총 데이터를 df 화하고, 
# 1. DAWON 폴더의 엑셀을 하나씩 불러와서 def로 집어 넣는다. 
# 2. def 안에서, ADDDATETIME이 오늘인거 날리고 (택스트로 변환 해야할거임)
# 3. 기존 총 데이터의 날짜들을 리스트로 만들고
# 4. 새로운 df 의 ADDDATETIME이 열의 값이, 위의 리스트에 있으면 날려버린다. 
# 5. 이때 새로운 df가 다 지워져 버리면 어떻게 되는지 모르겠음. 
# 6. 정리된 새로운 df를, 기존 데이터와 합치고, 
# 7. 업데이트 된 기존 데이터를 반환한다. 


excel_file_path = 'INV_master.xlsx'
# 엑셀 파일이 열려 있는지 확인하고, 열려 있다면 닫을 것을 안내합니다.
if os.path.exists(excel_file_path):
    messagebox.showinfo("파일닫기","마스터 파일을 닫아주세요. 제발요")



# 현재 날짜를 가져옵니다.
current_date = datetime.now()
# 날짜를 일반 문자 형식으로 포맷합니다.
formatted_date = current_date.strftime('%Y-%m-%d')



def update_dawon_data(df_dawon_all, df_new):
    lll=[]
    # 마지막 행 삭제
    df_new = df_new.iloc[:-1]

    #?--------------------------------------------------------
    #? 날짜+시간 형식을 날짜만 있는걸로 바꾸기
    #?--------------------------------------------------------
    # 날짜형식으로 변환
    df_dawon_all.loc[:,'ADDDATETIME'] = pd.to_datetime(df_dawon_all['ADDDATETIME']).dt.date
    df_new.loc[:,'ADDDATETIME'] = pd.to_datetime(df_new['ADDDATETIME']).dt.date
    #?--------------------------------------------------------

    # 가져온 all df 의 작업일자들을 리스트로 변환
    lll = df_new['ADDDATETIME'].unique().tolist()

    # 새로운 df에서, 기존 all df 와 작업일자가 겹치는 것들 삭제
    df_dawon_all = df_dawon_all[~df_dawon_all['ADDDATETIME'].isin(lll)]
    
    # 두 df 합치기
    df_dawon_all = pd.concat([df_dawon_all,df_new], ignore_index=True)

    return df_dawon_all


# 다원 올데이터 엑셀 파일 경로
excel_file_path = 'DAWON_all.xlsx'
# 엑셀 파일을 데이터프레임으로 불러오기
df_dawon_all = pd.read_excel(excel_file_path)

# 폴더 경로
folder_path = 'DAWON_출고'

# 폴더 내의 모든 엑셀 파일에 대해 반복
for filename in os.listdir(folder_path):
    #print(filename)
    if filename.endswith(('.xlsx', '.xls')):
        # 엑셀 파일 경로
        excel_file = os.path.join(folder_path, filename)
        # 엑셀 파일 불러오기
        df_new = pd.read_excel(excel_file)
        df_dawon_all = update_dawon_data(df_dawon_all, df_new) #? <--------- def는 여기 들어감

# 저장할 엑셀 파일 경로
excel_file_path = 'DAWON_all.xlsx'

# DataFrame을 엑셀 파일로 저장
df_dawon_all.to_excel(excel_file_path, index=False)  # index=False로 설정하여 인덱스를 엑셀 파일에 저장하지 않습니다.



#?----------------------------------------------------------------
#? 마스터 엑셀로 데이터 추출해서 이동
#?----------------------------------------------------------------
# 엑셀 파일 경로 설정
excel_file_path = 'INV_master.xlsx'
target_sheet_name = '출고_DAWON'  # 덮어쓸 시트의 이름

column_name = '출고완료일'
# 연-월-일 형식의 데이터를 연-월-일로 변경
df_dawon_all[column_name] = pd.to_datetime(df_dawon_all[column_name]).dt.date


# 복사할 열들 선택
columns_to_copy = ['출고완료일', 'ITEMGROUP', '품목코드(구성품)', '출고완료', '몰명']  # 복사할 열들의 이름

# 엑셀 파일에서 해당 시트를 불러옵니다.
with pd.ExcelWriter(excel_file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
    # 기존 엑셀 파일에 새로운 시트 추가
    df_dawon_all[columns_to_copy].to_excel(writer, sheet_name=target_sheet_name, index=False)






from openpyxl import load_workbook
from datetime import datetime

# 엑셀 파일을 열어서 서식을 변경합니다.
wb = load_workbook(excel_file_path)
ws = wb[target_sheet_name]

# "특정열1" 열의 서식을 "날짜"로 변경합니다.
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        # 셀의 값이 문자열인 경우에만 변환 수행
        if isinstance(cell.value, str):
            # yyyy-mm-dd 형식으로 되어있는 문자열을 파싱하여 년, 월, 일로 분리
            year, month, day = map(int, cell.value.split('-'))
            # 년, 월, 일을 "날짜" 형식으로 변환하여 셀에 쓰기
            cell.value = datetime(year, month, day).date()

# 변경된 내용을 저장합니다.
wb.save(excel_file_path)



#!----------------------------------------------------------------
#! 온스용 제작 
#!----------------------------------------------------------------
ons_excel_file_path = 'INV_master_ONNURISTORE.xlsx'

# 선택할 값들의 리스트
target_values = ['니심 ', '란시노 ', '아이로 ','에티튜드','조아써 ','웰라 ']

# 각 값에 대한 불리언 조건을 리스트 내포를 사용하여 생성
combined_condition = df_dawon_all['ITEMGROUP'].isin(target_values)

# 조건에 해당하는 행들만 선택
selected_rows = df_dawon_all[combined_condition].copy()



# 엑셀 파일에서 해당 시트를 불러옵니다.
with pd.ExcelWriter(ons_excel_file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
    # 기존 엑셀 파일에 새로운 시트 추가
    selected_rows[columns_to_copy].to_excel(writer, sheet_name=target_sheet_name, index=False)



# 엑셀 파일을 열어서 서식을 변경합니다.
wb = load_workbook(ons_excel_file_path)
ws = wb[target_sheet_name]

# "특정열1" 열의 서식을 "날짜"로 변경합니다.
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        # 셀의 값이 문자열인 경우에만 변환 수행
        if isinstance(cell.value, str):
            # yyyy-mm-dd 형식으로 되어있는 문자열을 파싱하여 년, 월, 일로 분리
            year, month, day = map(int, cell.value.split('-'))
            # 년, 월, 일을 "날짜" 형식으로 변환하여 셀에 쓰기
            cell.value = datetime(year, month, day).date()

# 변경된 내용을 저장합니다.
wb.save(ons_excel_file_path)